#!/usr/bin/env python3
"""
Actualiza los data/*.csv leyendo DIRECTO los archivos oficiales de INDEC e
IDECBA -- sin pasar por tu Excel personal.

Uso:
    python3 scripts/actualizar_desde_fuentes_oficiales.py \
        --indec sh_ipc_MM_AA.xls \
        --idecba-aperturas IPCBA_base_2021100-Principales_aperturas_indices.xlsx \
        --idecba-bs-svcios IPCBA_base_2021100-Evol_gral_bs_svcios.xlsx \
        --idecba-estac-reg-resto IPCBA_base_2021100-Evol_gral_estac_reg_resto.xlsx

Todos los flags son opcionales -- pasá solo los archivos que tengas nuevos.

Cómo funciona (empalme automático):
  Para cada serie, busca el último mes que YA tenés en data/<key>.csv y el
  mismo mes en el archivo fuente. Calcula el factor de escala entre ambos
  (por columna) y lo aplica a los meses del archivo fuente POSTERIORES a tu
  último dato, para agregarlos ya empalmados a la cadena histórica -- el
  mismo criterio que se usó a mano en todo este proyecto. Si el factor de
  escala no da ~1.0 en el mes de solape, es señal de que el archivo fuente
  no es continuación directa de tu serie (base distinta sin avisar, columnas
  reordenadas, etc.) y el script AVISA en vez de empalmar a ciegas.

Fuentes -> series que actualiza:
  --indec                  -> gba_sl_ipcba_gba_nacional, gba_slj_ipcba_gba_nacional
                               (bloque "Total nacional"), gba_sl_ipcba_gba (bloque "Región GBA")
  --idecba-aperturas        -> gba_sl_ipcba13 (Nivel General + 13 divisiones)
  --idecba-bs-svcios        -> gba_sl_ipcba13 (Bienes, Servicios)
  --idecba-estac-reg-resto  -> gba_sl_ipcba13 (Estacionales, Regulados, Resto IPCBA)

Nota: IDECBA ya no publica la apertura en 12 divisiones (gba_sl_ipcba12) --
la base 2021=100 solo tiene la de 13. Esa serie queda como referencia
histórica; avisame si querés que la derive de la de 13 divisiones o que la
demos de baja.

gba_sl_ipcba_gba_nacional_nacional2021 no se toca acá -- esa va con el
archivo "tipo INDEC" que armás vos para tu estimación ENGHo propia (correr
ese con el otro flag --indec cuando lo tengas, apunta al mismo lector).
"""
import sys
import csv
import argparse
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _lib import DATA_DIR, build_bundle

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ---------------------------------------------------------------------------
# utilidades comunes
# ---------------------------------------------------------------------------

def load_csv(key):
    path = DATA_DIR / f"{key}.csv"
    with open(path, encoding="utf-8") as f:
        r = csv.reader(f)
        header = next(r)
        rows = {row[0]: row[1:] for row in r}
    return header, rows


def save_csv(key, header, rows):
    path = DATA_DIR / f"{key}.csv"
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for ym in sorted(rows):
            w.writerow([ym] + rows[ym])


def splice_and_merge(key, new_data, label):
    """new_data: {ym: {col_name: valor}}. Empalma contra data/<key>.csv."""
    header, rows = load_csv(key)
    cols = header[1:]
    existing_last = max(rows.keys())

    if existing_last not in new_data:
        print(f"  [!] {key}: tu último dato ({existing_last}) no está en '{label}' "
              f"-- no puedo calibrar el empalme, no toco nada.")
        return None

    overlap = new_data[existing_last]
    scale = {}
    for c in cols:
        old_v = rows[existing_last][cols.index(c)]
        new_v = overlap.get(c)
        if old_v in (None, "") or new_v in (None, ""):
            continue
        old_v = float(old_v)
        if abs(new_v) < 1e-12:
            continue
        scale[c] = old_v / new_v

    bad = {c: s for c, s in scale.items() if not (0.9 < s < 1.1)}
    if bad:
        print(f"  [!] {key}: el factor de empalme contra '{label}' se aleja de 1.0 "
              f"en {list(bad)[:3]}{'...' if len(bad) > 3 else ''} -- no toco nada, "
              f"revisalo a mano (¿archivo de otra base/región?).")
        return None

    new_months = sorted(ym for ym in new_data if ym > existing_last)
    if not new_months:
        print(f"  [=] {key}: sin novedades en '{label}' más allá de {existing_last}.")
        return None

    for ym in new_months:
        row = []
        for c in cols:
            v = new_data[ym].get(c)
            if v is None or c not in scale:
                row.append("")
            else:
                row.append(v * scale[c])
        rows[ym] = row

    save_csv(key, header, rows)
    print(f"  [OK] {key}: agregué {len(new_months)} mes(es) nuevo(s) desde '{label}' "
          f"({new_months[0]} a {new_months[-1]}), empalmados (factor promedio "
          f"{sum(scale.values())/len(scale):.4f}).")
    return new_months[0], new_months[-1]


# ---------------------------------------------------------------------------
# INDEC -- "Índices IPC Cobertura Nacional" (sh_ipc_MM_AA.xls, formato oficial)
# ---------------------------------------------------------------------------

# offsets de fila relativos al inicio de cada bloque regional ("Total nacional",
# "Región GBA", ...), calibrados a mano contra el layout oficial de INDEC.
# El orden coincide exactamente con el de nuestras columnas -- por eso se
# puede mapear por POSICIÓN, sin depender de que el nombre coincida letra
# por letra (INDEC nombra distinto un par de rubros, ej. "Comunicación" en
# vez de "Comunicaciones").
_INDEC_OFFSETS = [4,5,6,7,8,9,10,11,12,13,14,15,16,19,20,21,24,25]
_BLOCK_STRIDE = 30  # filas entre el inicio de un bloque regional y el siguiente


def read_indec_xls(path):
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name("Índices IPC Cobertura Nacional")

    def xldate(n):
        return datetime(1899, 12, 30) + timedelta(days=int(n))

    def read_block(start_row):
        dates_row = ws.row_values(start_row)
        cols = {}
        for ci in range(1, ws.ncols):
            v = dates_row[ci]
            if isinstance(v, (int, float)) and v > 0:
                ym = f"{xldate(v).year:04d}-{xldate(v).month:02d}"
                vals = {}
                for oi, off in enumerate(_INDEC_OFFSETS):
                    row_idx = start_row + off
                    if row_idx < ws.nrows:
                        vals[oi] = ws.cell_value(row_idx, ci)
                cols[ym] = vals
        return cols

    def find_row(label):
        for r in range(ws.nrows):
            if str(ws.cell_value(r, 0)).strip() == label:
                return r
        return None

    nacional_start = find_row("Total nacional")
    gba_start = find_row("Región GBA")
    nacional = read_block(nacional_start) if nacional_start is not None else {}
    gba = read_block(gba_start) if gba_start is not None else {}
    return nacional, gba


def read_indec_xlsx(path):
    """Por si algún día INDEC publica en .xlsx en vez de .xls, mismo layout."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Índices IPC Cobertura Nacional"]
    rows = list(ws.iter_rows(values_only=True))

    def read_block(start_row):
        dates_row = rows[start_row]
        cols = {}
        for ci in range(1, len(dates_row)):
            v = dates_row[ci]
            if isinstance(v, datetime):
                ym = f"{v.year:04d}-{v.month:02d}"
                vals = {}
                for oi, off in enumerate(_INDEC_OFFSETS):
                    r = start_row + off
                    if r < len(rows):
                        vals[oi] = rows[r][ci]
                cols[ym] = vals
        return cols

    def find_row(label):
        for i, row in enumerate(rows):
            if row and str(row[0]).strip() == label:
                return i
        return None

    nacional_start = find_row("Total nacional")
    gba_start = find_row("Región GBA")
    nacional = read_block(nacional_start) if nacional_start is not None else {}
    gba = read_block(gba_start) if gba_start is not None else {}
    return nacional, gba


def by_position(block, col_names):
    """{ym: {0:v,1:v,...}} -> {ym: {col_name: v}} usando el orden de col_names."""
    out = {}
    for ym, vals in block.items():
        out[ym] = {col_names[i]: v for i, v in vals.items() if i < len(col_names)}
    return out


def update_from_indec(path):
    print(f"\nLeyendo INDEC: {path.name}")
    if path.suffix.lower() == ".xls":
        nacional_raw, gba_raw = read_indec_xls(str(path))
    else:
        nacional_raw, gba_raw = read_indec_xlsx(str(path))

    for key in ("gba_sl_ipcba_gba_nacional", "gba_slj_ipcba_gba_nacional"):
        header, _ = load_csv(key)
        nacional = by_position(nacional_raw, header[1:])
        splice_and_merge(key, nacional, "INDEC Total nacional")

    header, _ = load_csv("gba_sl_ipcba_gba")
    gba = by_position(gba_raw, header[1:])
    splice_and_merge("gba_sl_ipcba_gba", gba, "INDEC Región GBA")


# ---------------------------------------------------------------------------
# IDECBA -- archivos "IPCBA base 2021=100"
# ---------------------------------------------------------------------------

def _read_dated_rows_block(path, sheet=None):
    """Lee archivos con fechas en FILAS (Nivel_general_empalme, Bs_svcios_*,
    Estac_reg_resto_*, Evol_gral_*): header en la fila con nombres de columna,
    datos abajo. Si el header se repite (bloque de variación % al lado), se
    queda solo con el primer bloque (los niveles, no las variaciones)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    header_row = None
    for i, row in enumerate(rows):
        if row and isinstance(row[0], type(None)) and any(
                isinstance(c, str) and "nivel general" in c.lower() for c in row if c):
            header_row = i
            break
    if header_row is None:
        raise ValueError(f"No encontré la fila de encabezados en {path}")

    header = rows[header_row]
    # ancho del primer bloque: hasta que el nombre de columna se repita
    seen = set()
    width = 1
    for c in header[1:]:
        name = str(c).strip() if c else None
        if name in seen:
            break
        if name:
            seen.add(name)
        width += 1
    col_names = [str(c).strip() for c in header[1:width]]

    out = {}
    for row in rows[header_row + 1:]:
        d = row[0]
        if not isinstance(d, datetime):
            continue
        ym = f"{d.year:04d}-{d.month:02d}"
        out[ym] = {col_names[i]: row[1 + i] for i in range(len(col_names))
                   if row[1 + i] is not None}
    return out


def _read_aperturas_transposed(path, sheet=None, wanted=None):
    """Lee Principales_aperturas_*: rubros en FILAS (col A), fechas en
    columnas. `wanted` es la lista de nombres de rubro a extraer (se ignoran
    las subaperturas más granulares que no usamos)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    date_row = None
    for i, row in enumerate(rows):
        if row and any(isinstance(c, datetime) for c in row):
            date_row = row
            break
    if date_row is None:
        raise ValueError(f"No encontré la fila de fechas en {path}")

    date_cols = {ci: c for ci, c in enumerate(date_row) if isinstance(c, datetime)}

    out = {f"{d.year:04d}-{d.month:02d}": {} for d in date_cols.values()}
    wanted_norm = {w.strip().lower(): w for w in (wanted or [])}
    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if wanted and name.lower() not in wanted_norm:
            continue
        target_name = wanted_norm.get(name.lower(), name)
        for ci, d in date_cols.items():
            ym = f"{d.year:04d}-{d.month:02d}"
            v = row[ci] if ci < len(row) else None
            if v is not None:
                out[ym][target_name] = v
    return out


def update_from_idecba(aperturas=None, bs_svcios=None, estac_reg_resto=None):
    key = "gba_sl_ipcba13"
    header, _ = load_csv(key)
    cols = header[1:]

    merged = {}

    if aperturas:
        print(f"\nLeyendo IDECBA aperturas: {aperturas.name}")
        wanted = [c for c in cols if c not in ("Bienes", "Servicios", "Estacionales ", "Regulados", "Resto IPCBA")]
        data = _read_aperturas_transposed(aperturas, wanted=wanted)
        for ym, vals in data.items():
            merged.setdefault(ym, {}).update(vals)

    if bs_svcios:
        print(f"Leyendo IDECBA bienes/servicios: {bs_svcios.name}")
        data = _read_dated_rows_block(bs_svcios)
        for ym, vals in data.items():
            merged.setdefault(ym, {}).update(
                {k: v for k, v in vals.items() if k in ("Bienes", "Servicios")})

    if estac_reg_resto:
        print(f"Leyendo IDECBA estacionales/regulados/resto: {estac_reg_resto.name}")
        data = _read_dated_rows_block(estac_reg_resto)
        for ym, vals in data.items():
            for k, v in vals.items():
                kk = k if k in cols else (k + " " if (k + " ") in cols else k)
                merged.setdefault(ym, {})[kk] = v

    if not merged:
        return
    splice_and_merge(key, merged, "IDECBA base 2021=100")

    print("\n  [i] gba_sl_ipcba12 (12 divisiones) no se actualizó -- IDECBA ya no "
          "publica esa apertura en la base 2021=100. Queda como estaba.")


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument("--indec")
    ap.add_argument("--idecba-aperturas")
    ap.add_argument("--idecba-bs-svcios")
    ap.add_argument("--idecba-estac-reg-resto")
    ap.add_argument("-h", "--help", action="store_true")
    args = ap.parse_args()

    if args.help or not any([args.indec, args.idecba_aperturas, args.idecba_bs_svcios,
                              args.idecba_estac_reg_resto]):
        print(__doc__)
        sys.exit(0)

    if args.indec:
        update_from_indec(Path(args.indec))

    if args.idecba_aperturas or args.idecba_bs_svcios or args.idecba_estac_reg_resto:
        update_from_idecba(
            Path(args.idecba_aperturas) if args.idecba_aperturas else None,
            Path(args.idecba_bs_svcios) if args.idecba_bs_svcios else None,
            Path(args.idecba_estac_reg_resto) if args.idecba_estac_reg_resto else None,
        )

    build_bundle()


if __name__ == "__main__":
    main()
