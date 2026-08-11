#!/usr/bin/env python3
"""
Actualiza los data/*.csv de la calculadora leyendo DIRECTO de tu Excel maestro
("Inflación.xlsx"), sin necesidad de exportar cada hoja a .txt a mano.

Uso:
    python3 scripts/actualizar_desde_excel.py /ruta/a/Inflación.xlsx

Qué hace:
  1. Abre el libro y ubica, para cada una de las 6 series, la hoja que le
     corresponde (ver SHEET_MAP más abajo).
  2. Cada hoja tiene el mismo bloque de columnas (Mes, Nivel general, rubro
     por rubro...) repetido varias veces al costado (variación % mensual,
     nivel encadenado, un índice rebasado a 100, etc.). El script CALIBRA
     solo cuál repetición es el nivel encadenado real: prueba cada bloque
     candidato contra el CSV que ya tenés en data/ y se queda con el que
     matchea sin error en todos los meses que se solapan.
  3. Vuelca ese bloque (todas las columnas, no solo "Nivel general") a
     data/<key>.csv, igual que el otro script.
  4. Regenera data/bundle.js.
  5. Te tira las líneas para pegar en data/series.js.

Si la calibración no encuentra un bloque que matchee sin error para alguna
serie, el script AVISA y no toca ese CSV — mejor que pisarlo con algo mal
alineado. Lo más probable en ese caso es que hayas reordenado columnas en esa
hoja del Excel; revisá SHEET_MAP / BLOCK_WIDTH abajo.
"""
import sys
import csv
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _lib import DATA_DIR, build_bundle, print_series_js_snippet

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl: pip install openpyxl --break-system-packages")
    sys.exit(1)

# Nunca tomamos el mes calendario actual ni posteriores: un mes en curso no
# puede tener dato oficial publicado todavía, así que si aparece algo cargado
# ahí (proyección/estimación propia mezclada en la misma hoja) lo cortamos
# antes de que se cuele en el sitio como si fuera dato real.
HOY = datetime.now()
MES_ACTUAL_YM = f"{HOY.year:04d}-{HOY.month:02d}"

# key interna -> nombre de la hoja en el Excel que la contiene
SHEET_MAP = {
    "gba_sl_ipcba_gba_nacional": "Pond2",
    "gba_slj_ipcba_gba_nacional": "Pond",
    "gba_sl_ipcba_gba": "GBA-SL-IPCBA-GBA",
    "gba_sl_ipcba_gba_nacional_nacional2021": "Pond2021",
    "gba_sl_ipcba12": "IPCBA12",
    "gba_sl_ipcba13": "IPCBA",
}

# cantidad de columnas de rubro (sin contar "Mes") que tiene el bloque bueno
# en cada hoja -- ver data/series.js "columns" de cada serie.
BLOCK_WIDTH = {
    "gba_sl_ipcba_gba_nacional": 18,
    "gba_slj_ipcba_gba_nacional": 18,
    "gba_sl_ipcba_gba": 18,
    "gba_sl_ipcba_gba_nacional_nacional2021": 18,
    "gba_sl_ipcba12": 18,
    "gba_sl_ipcba13": 19,
}

NIVEL_NAMES = {"nivel general"}

# algunas hojas comparten "Nivel General" con otra serie ya validada acá adentro
# (después de 2017 el San Luis-Jujuy es igual al San Luis; 12 y 13 divisiones
# comparten el mismo Nivel General de IPCBA). Se usa para detectar y cortar
# meses que en el Excel están completados con una proyección/estimación propia
# en vez de dato realmente publicado -- si la serie hija tiene meses que la
# serie madre todavía no tiene, se asume que son estimación y se recortan.
SIBLING_MAP = {
    "gba_slj_ipcba_gba_nacional": "gba_sl_ipcba_gba_nacional",
    "gba_sl_ipcba13": "gba_sl_ipcba12",
}
# procesar madres antes que hijas, para que el recorte use el dato recién
# calculado en esta misma corrida y no un CSV viejo en disco
_PROCESS_ORDER = [k for k in SHEET_MAP if k not in SIBLING_MAP] + list(SIBLING_MAP)


def load_ref(key):
    """Serie 'Nivel general' que ya tenemos validada, para calibrar contra ella."""
    path = DATA_DIR / f"{key}.csv"
    if not path.exists():
        return {}
    ref = {}
    with open(path, encoding="utf-8") as f:
        r = csv.reader(f)
        next(r)
        for row in r:
            if row and row[1] != "":
                ref[row[0]] = float(row[1])
    return ref


def find_block_start(rows, header, ref):
    """Encuentra el índice de columna donde arranca el bloque 'Nivel general'
    real, probando cada candidato contra la serie de referencia ya validada."""
    starts = [i for i, h in enumerate(header)
              if isinstance(h, str) and h.strip().lower() in NIVEL_NAMES]
    for s in starts:
        checked = 0
        ok = True
        for row in rows:
            d = row[0]
            if d is None:
                continue
            ym = f"{d.year:04d}-{d.month:02d}"
            if ym in ref:
                v = row[s]
                try:
                    if v is None or abs(float(v) - ref[ym]) > 1e-6 * max(1, abs(ref[ym])):
                        ok = False
                        break
                except (TypeError, ValueError):
                    ok = False
                    break
                checked += 1
        if ok and checked > 0:
            return s
    return None


def convert_sheet(wb, key, results_so_far):
    sheet_name = SHEET_MAP[key]
    if sheet_name not in wb.sheetnames:
        print(f"  [!] La hoja '{sheet_name}' (para {key}) no está en este Excel — la salteo.")
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    ref = load_ref(key)
    if not ref:
        print(f"  [!] No tengo data/{key}.csv para calibrar contra — la salteo. "
              f"(hace falta al menos una vez el CSV ya generado por el otro script)")
        return None

    start = find_block_start(data_rows, header, ref)
    if start is None:
        print(f"  [!] {key}: no encontré un bloque en '{sheet_name}' que matchee "
              f"la referencia sin error. No toco el CSV — revisá a mano.")
        return None

    width = BLOCK_WIDTH[key]
    cols = [str(header[start + i]).strip() if header[start + i] is not None else ""
            for i in range(1, width)]
    # la primera columna del bloque es "Nivel general" / "Nivel General" tal
    # como está escrita en el Excel (respeta may/min y espacios originales)
    cols = [str(header[start]).strip()] + cols

    out_rows = []
    truncated = 0
    for row in data_rows:
        d = row[0]
        if d is None:
            continue
        ym = f"{d.year:04d}-{d.month:02d}"
        if ym >= MES_ACTUAL_YM:
            truncated += 1
            continue
        vals = [row[start + i] if start + i < len(row) else None for i in range(width)]
        out_rows.append((ym, vals))

    if not out_rows:
        print(f"  [!] {key}: la hoja no tiene filas de fecha (anteriores a {MES_ACTUAL_YM}). La salteo.")
        return None

    # recorte contra la serie madre, si corresponde -- ver SIBLING_MAP arriba.
    # cubre el caso en que la hoja tiene proyección/estimación propia cargada
    # en meses ya pasados (no solo en el mes en curso), que el corte por fecha
    # de arriba no detecta.
    sibling = SIBLING_MAP.get(key)
    if sibling and sibling in results_so_far:
        sibling_last = results_so_far[sibling][1]  # (first_ym, last_ym)
        before = len(out_rows)
        excess = [ym for ym, _ in out_rows if ym > sibling_last]
        out_rows = [(ym, vals) for ym, vals in out_rows if ym <= sibling_last]
        if excess:
            print(f"  [i] {key}: recorté {len(excess)} mes(es) más allá de {sibling_last} "
                  f"({excess[0]} a {excess[-1]}) por no tener respaldo en '{sibling}' "
                  f"-- asumidos estimación/proyección propia, no dato publicado.")

    if not out_rows:
        print(f"  [!] {key}: no quedaron filas después del recorte contra '{sibling}'. La salteo.")
        return None

    out_path = DATA_DIR / f"{key}.csv"
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["ym"] + cols)
        for ym, vals in out_rows:
            w.writerow([ym] + ["" if v is None else v for v in vals])

    # último mes con dato real en Nivel general (no solo con fecha cargada)
    last_ym = None
    first_ym = out_rows[0][0]
    for ym, vals in out_rows:
        if vals[0] is not None:
            last_ym = ym
    print(f"  [OK] {sheet_name} (col {start+1}) -> data/{key}.csv  "
          f"({len(out_rows)} filas, {first_ym} a {last_ym})"
          + (f"  [{truncated} fila(s) del mes actual/futuras descartadas]" if truncated else ""))
    return first_ym, last_ym


def main(argv):
    if len(argv) != 1:
        print(__doc__)
        sys.exit(1)

    xlsx_path = Path(argv[0])
    if not xlsx_path.exists():
        print(f"No encuentro {xlsx_path}")
        sys.exit(1)

    print(f"Abriendo {xlsx_path.name}... (puede tardar unos segundos, es un libro grande)\n")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    results = {}
    for key in _PROCESS_ORDER:
        r = convert_sheet(wb, key, results)
        if r:
            results[key] = r

    if not results:
        print("\nNo se actualizó ninguna serie.")
        return

    build_bundle()
    print_series_js_snippet([(k, v[0], v[1]) for k, v in results.items()])


if __name__ == "__main__":
    main(sys.argv[1:])
